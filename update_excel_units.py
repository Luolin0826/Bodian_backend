#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修改25国网南网录取.xlsx文件中的录取单位名称
为指定的省份添加"电网"后缀
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def update_units():
    """更新Excel文件中的录取单位名称"""
    
    # 第一批需要添加"电网"的单位
    first_batch_units = [
        "安徽", "北京", "福建", "甘肃", "河北", "河南", "黑龙江", "湖北", "湖南", 
        "吉林", "冀北", "江苏", "江西", "辽宁", "蒙东", "宁夏", "青海", "山东", 
        "山西", "陕西", "上海", "四川", "天津", "西藏", "新疆", "浙江", "重庆"
    ]
    
    # 第二批需要添加"电网"的单位
    second_batch_units = [
        "宁夏", "青海", "山东", "山西", "陕西", "上海", "四川", 
        "天津", "西藏", "新疆", "浙江", "重庆"
    ]
    
    # 读取原文件
    file_path = '/workspace/25国网南网录取.xlsx'
    
    # 使用openpyxl保持格式
    wb = load_workbook(file_path)
    
    # 处理第一批
    ws1 = wb['一批']
    
    # 找到"上岸电网省份"列的位置
    header_row = 1
    unit_col_idx = None
    for col_idx, cell in enumerate(ws1[header_row], 1):
        if cell.value == "上岸电网省份":
            unit_col_idx = col_idx
            break
    
    if unit_col_idx:
        print(f"找到第一批'上岸电网省份'列在第{unit_col_idx}列")
        
        # 遍历所有数据行
        for row_idx in range(2, ws1.max_row + 1):
            cell = ws1.cell(row=row_idx, column=unit_col_idx)
            if cell.value and str(cell.value).strip() in first_batch_units:
                old_value = str(cell.value).strip()
                new_value = old_value + "电网"
                cell.value = new_value
                print(f"第一批: {old_value} -> {new_value}")
    
    # 处理第二批
    ws2 = wb['二批']
    
    # 找到"上岸电网省份"列的位置
    unit_col_idx = None
    for col_idx, cell in enumerate(ws2[header_row], 1):
        if cell.value == "上岸电网省份":
            unit_col_idx = col_idx
            break
    
    if unit_col_idx:
        print(f"找到第二批'上岸电网省份'列在第{unit_col_idx}列")
        
        # 遍历所有数据行
        for row_idx in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=row_idx, column=unit_col_idx)
            if cell.value and str(cell.value).strip() in second_batch_units:
                old_value = str(cell.value).strip()
                new_value = old_value + "电网"
                cell.value = new_value
                print(f"第二批: {old_value} -> {new_value}")
    
    # 保存修改后的文件
    output_path = '/workspace/25国网南网录取_updated.xlsx'
    wb.save(output_path)
    print(f"\n文件已保存到: {output_path}")
    
    return output_path

if __name__ == "__main__":
    update_units()